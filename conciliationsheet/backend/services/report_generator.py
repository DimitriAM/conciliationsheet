from datetime import datetime
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from database.db import get_connection


def _fetch_partidas(cuenta_id: Optional[int] = None,
                    tipo: Optional[str] = None,
                    estado: Optional[str] = None,
                    fecha_desde: Optional[str] = None,
                    fecha_hasta: Optional[str] = None) -> list[dict]:
    conn = get_connection()
    condiciones = []
    params = []
    if cuenta_id:
        condiciones.append("pc.cuenta_id = ?")
        params.append(cuenta_id)
    if tipo:
        condiciones.append("pc.tipo = ?")
        params.append(tipo)
    if estado:
        condiciones.append("pc.estado = ?")
        params.append(estado)
    if fecha_desde:
        condiciones.append("pc.fecha >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        condiciones.append("pc.fecha <= ?")
        params.append(fecha_hasta)
    where = "WHERE " + " AND ".join(condiciones) if condiciones else ""
    rows = conn.execute(
        f"SELECT * FROM partidas_conciliatorias pc {where} ORDER BY pc.fecha DESC", params
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _get_resumen_partidas(cuenta_id: Optional[int] = None,
                          tipo: Optional[str] = None,
                          estado: Optional[str] = None,
                          fecha_desde: Optional[str] = None,
                          fecha_hasta: Optional[str] = None) -> list[dict]:
    conn = get_connection()
    condiciones = []
    params = []
    if cuenta_id:
        condiciones.append("pc.cuenta_id = ?")
        params.append(cuenta_id)
    if tipo:
        condiciones.append("pc.tipo = ?")
        params.append(tipo)
    if estado:
        condiciones.append("pc.estado = ?")
        params.append(estado)
    if fecha_desde:
        condiciones.append("pc.fecha >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        condiciones.append("pc.fecha <= ?")
        params.append(fecha_hasta)
    where = "WHERE " + " AND ".join(condiciones) if condiciones else ""
    rows = conn.execute(
        f"""SELECT pc.tipo, COUNT(*) as cantidad,
                   COALESCE(SUM(pc.debe),0) as total_debe,
                   COALESCE(SUM(pc.haber),0) as total_haber
            FROM partidas_conciliatorias pc {where}
            GROUP BY pc.tipo""", params
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def generar_excel_partidas(cuenta_id=None, tipo=None, estado=None,
                           fecha_desde=None, fecha_hasta=None) -> BytesIO:
    datos = _fetch_partidas(cuenta_id, tipo, estado, fecha_desde, fecha_hasta)
    resumen = _get_resumen_partidas(cuenta_id, tipo, estado, fecha_desde, fecha_hasta)

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Resumen"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws_sum.cell(1, 1, "Reporte de Partidas Conciliatorias").font = Font(bold=True, size=14)
    ws_sum.cell(2, 1, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10)
    ws_sum.cell(4, 1, "Tipo").font = header_font
    ws_sum.cell(4, 1).fill = header_fill
    ws_sum.cell(4, 2, "Cantidad").font = header_font
    ws_sum.cell(4, 2).fill = header_fill
    ws_sum.cell(4, 3, "Total Debe").font = header_font
    ws_sum.cell(4, 3).fill = header_fill
    ws_sum.cell(4, 4, "Total Haber").font = header_font
    ws_sum.cell(4, 4).fill = header_fill

    for i, r in enumerate(resumen, 5):
        ws_sum.cell(i, 1, r["tipo"])
        ws_sum.cell(i, 2, r["cantidad"])
        ws_sum.cell(i, 3, round(r["total_debe"], 2))
        ws_sum.cell(i, 4, round(r["total_haber"], 2))

    for col in range(1, 5):
        cell = ws_sum.cell(4, col)
        cell.border = thin_border

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 15

    ws_det = wb.create_sheet("Detalle")
    headers = ["ID", "Fecha", "Descripcion", "Tipo", "Origen",
               "Debe", "Haber", "Saldo Afectado", "Estado", "Observaciones"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws_det.cell(1, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, d in enumerate(datos, 2):
        ws_det.cell(row_idx, 1, d["id"])
        ws_det.cell(row_idx, 2, d["fecha"])
        ws_det.cell(row_idx, 3, d["descripcion"])
        ws_det.cell(row_idx, 4, d["tipo"])
        ws_det.cell(row_idx, 5, d["origen"])
        ws_det.cell(row_idx, 6, round(d["debe"], 2))
        ws_det.cell(row_idx, 7, round(d["haber"], 2))
        ws_det.cell(row_idx, 8, d["saldo_afectado"])
        ws_det.cell(row_idx, 9, d["estado"])
        ws_det.cell(row_idx, 10, d["observaciones"] or "")

    for col_idx in range(1, len(headers) + 1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_pdf_partidas(cuenta_id=None, tipo=None, estado=None,
                         fecha_desde=None, fecha_hasta=None) -> BytesIO:
    datos = _fetch_partidas(cuenta_id, tipo, estado, fecha_desde, fecha_hasta)
    resumen = _get_resumen_partidas(cuenta_id, tipo, estado, fecha_desde, fecha_hasta)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Reporte de Partidas Conciliatorias", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 10)
    pdf.cell(0, 8, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Resumen", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(40, 7, "Tipo", border=1, align="C")
    pdf.cell(25, 7, "Cantidad", border=1, align="C")
    pdf.cell(30, 7, "Total Debe", border=1, align="C")
    pdf.cell(30, 7, "Total Haber", border=1, align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    for r in resumen:
        pdf.cell(40, 7, r["tipo"], border=1)
        pdf.cell(25, 7, str(r["cantidad"]), border=1, align="C")
        pdf.cell(30, 7, f"{r['total_debe']:.2f}", border=1, align="R")
        pdf.cell(30, 7, f"{r['total_haber']:.2f}", border=1, align="R",
                 new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Detalle de Partidas", new_x="LMARGIN", new_y="NEXT")

    headers = ["Fecha", "Descripcion", "Tipo", "Debe", "Haber", "Estado"]
    col_w = [22, 60, 25, 22, 22, 22]
    pdf.set_font("Helvetica", "B", 8)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 6, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for d in datos:
        row = [
            d["fecha"][:10] if d["fecha"] else "",
            d["descripcion"][:35] if d["descripcion"] else "",
            d["tipo"] or "",
            f"{d['debe']:.2f}",
            f"{d['haber']:.2f}",
            d["estado"] or "",
        ]
        for i, val in enumerate(row):
            pdf.cell(col_w[i], 5, val, border=1, align="C" if i > 0 else "L")
        pdf.ln()

    output = BytesIO()
    pdf.output(output)
    output.seek(0)
    return output
